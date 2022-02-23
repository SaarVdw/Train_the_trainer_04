import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ingeven csv
# ingeven padnaam naar csv

lijst = pd.read_csv("export - 2022-02-23T094438.836.csv", delimiter=';')
print(lijst)

#########################################################################

# Maak lijsten met ontbrekende velden
# 1. Selecteer ontbrekende velden via pandas
# 1.1 begindatum

# dentificeer wanneer begindatum 'leeg' is
ontbrekende_begindatum = pd.isna(lijst['vervaardiging.datum.begin'])
print(ontbrekende_begindatum)

# hou vanuit de lijst enkel de records over waarbij begindatum leeg = true was
ontbrekende_begindatum = lijst[ontbrekende_begindatum]
print(ontbrekende_begindatum)

# 1.2 vervaardiger

ontbrekende_vervaardiger = pd.isna(lijst['vervaardiger'])
print(ontbrekende_vervaardiger)

ontbrekende_vervaardiger = lijst[ontbrekende_vervaardiger]
print(ontbrekende_vervaardiger)

# 1.3 beschrijving

ontbrekende_beschrijving = pd.isna(lijst['beschrijving'])

ontbrekende_beschrijving = lijst[ontbrekende_beschrijving]
print(ontbrekende_beschrijving)

# 1.4 rol vervaardiger

ontbrekende_rol = pd.isna(lijst['vervaardiger.rol'])

ontbrekende_rol = lijst[ontbrekende_rol]
print(ontbrekende_rol)

# 1.5 instellingsnaam

ontbrekende_instelling = pd.isna(lijst['instelling.naam'])

ontbrekende_instelling = lijst[ontbrekende_instelling]
print(ontbrekende_instelling)

# 1.6 vervaardiging plaats

ontbrekende_plaats = pd.isna(lijst['vervaardiging.plaats'])

ontbrekende_plaats = lijst[ontbrekende_plaats]
print(ontbrekende_plaats)

# 1.7 einddatum

ontbrekende_einddatum = pd.isna(lijst['vervaardiging.datum.eind'])

ontbrekende_einddatum = lijst[ontbrekende_einddatum]
print(ontbrekende_einddatum)

# 1.8 titel

ontbrekende_titel = pd.isna(lijst['titel'])

ontbrekende_titel = lijst[ontbrekende_titel]
print(ontbrekende_titel)

####################################################################################

# B. maak lijsten met ontbrekende velden
# 2. Output (excel) maken met lijsten met ontbrekende velden (via openpyxl)

flore = Workbook()
flore.save(r'C:\Users\vandewsa\PycharmProjects\TTT4\output.xlsx')

# maak sheet (tabblad)
# begindatum
ws = flore.create_sheet("begindatum")
# zet dataframe (pandas) om naar rijen in het tabblad
rows = dataframe_to_rows(ontbrekende_begindatum, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

#beschrijving
ws = flore.create_sheet("beschrijving")
rows = dataframe_to_rows(ontbrekende_beschrijving, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

#einddatum
ws = flore.create_sheet("einddatum")
rows = dataframe_to_rows(ontbrekende_einddatum, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

#instelling
ws = flore.create_sheet("instelling")
rows = dataframe_to_rows(ontbrekende_instelling, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

#plaats vervaardiging
ws = flore.create_sheet("plaats vervaardiging")
rows = dataframe_to_rows(ontbrekende_plaats, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

#rol
ws = flore.create_sheet("rol vervaardiger")
rows = dataframe_to_rows(ontbrekende_rol, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

#titel
ws = flore.create_sheet("titel")
rows = dataframe_to_rows(ontbrekende_titel, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

#vervaardiger
ws = flore.create_sheet("vervaardiger")
rows = dataframe_to_rows(ontbrekende_vervaardiger, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# opslaan
flore.save(r'C:\Users\vandewsa\PycharmProjects\TTT4\output.xlsx')

#######################################################################################################################
# C. visualisatie van de data
# 1. tellen van de data (pandas)

# tel aantal keer objectnummer aanwezig is in lijst records zonder objectnamen (gezien objectnummer altijd aanwezig)
ontbrekende_begindata = ontbrekende_begindatum['objectnummer'].count()
print(ontbrekende_begindatum)

# doe hetzelfde voor de overige velden
ontbrekende_titels = ontbrekende_titel['objectnummer'].count()
ontbrekende_beschrijvingen = ontbrekende_beschrijving['objectnummer'].count()
ontbrekende_vervaardigers = ontbrekende_vervaardiger['objectnummer'].count()
ontbrekende_einddata = ontbrekende_einddatum['objectnummer'].count()
ontbrekende_rollen = ontbrekende_rol['objectnummer'].count()
ontbrekende_plaatsen = ontbrekende_plaats['objectnummer'].count()

#######################################################################################################################
# C. visualisatie van de data
# 2. weergeven van de data in grafiek in excel (openpyxl)

# 2.1 voeg de data toe in excel
# zet de data in een list
labels = ["titel", "beschrijving", "vervaardiger", "begindatum", "einddatum", "rol", "plaats vervaardiging"]
ontbrekende_data = [ontbrekende_titels, ontbrekende_beschrijvingen, ontbrekende_vervaardigers,
                    ontbrekende_begindata, ontbrekende_einddata, ontbrekende_rollen, ontbrekende_plaatsen]

# voeg de lijst toe aan de excel (openpyxl)

ws = flore.active
ws.title = 'Info'
# titel lijst
ws['A1'] = "Ontbrekende velden"
flore.save(r'C:\Users\vandewsa\PycharmProjects\TTT4\output.xlsx')

# de labels
rij1 = 2
for label in labels:
    ws.cell(row=rij1, column=1).value = label
    rij1 += 1
flore.save(r'C:\Users\vandewsa\PycharmProjects\TTT4\output.xlsx')

# de waardes
rij2 = 2
for veld in ontbrekende_data:
    ws.cell(row=rij2, column=2).value = veld
    rij2 += 1
flore.save(r'C:\Users\vandewsa\PycharmProjects\TTT4\output.xlsx')

# 2.2 maak de grafiek in excel

from openpyxl.chart import BarChart3D, Reference

# selecteer de data
data = Reference(ws, min_col=2, min_row=2, max_col=2, max_row=6)

# selecteer de labels
labels = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=6)

# creeer grafiek
chart = BarChart3D()

# voeg grafiek toe aan excel
ws.add_chart(chart, "E2")
flore.save(r'C:\Users\vandewsa\PycharmProjects\TTT4\output.xlsx')

# voeg titel toe
chart.title = 'Ontbrekende data'

# voeg namen x en y as toe
chart.y_axis.title = 'aantal'
chart.x_axis.title = 'velden'

# voeg de data & labels toe
chart.add_data(data)
chart.set_categories(labels)

flore.save(r'C:\Users\vandewsa\PycharmProjects\TTT4\output.xlsx')